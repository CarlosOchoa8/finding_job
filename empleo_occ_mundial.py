import difflib
import os
import re
import socket
import time

import openpyxl
import openpyxl.styles
import openpyxl.utils
import openpyxl.worksheet
import openpyxl.worksheet.dimensions
import pandas as pd
import xlsxwriter
from openpyxl.styles import Font, PatternFill
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

service = Service(ChromeDriverManager().install())
option = webdriver.ChromeOptions()

# option.add_argument("--headless") # ejecucion sin que se abra en ventana
option.add_argument("--start-maximized") # tama;o de pantalla estandar

# # creacion de navegador
browser = webdriver.Chrome(options=option, service=service)

# Peticion a sitio
browser.get("https://www.occ.com.mx")

# Palabras clave
POSITION = "python"
LOCATION = "Mexico"

# Busqueda campos puesto/estado
position_input = browser.find_element(By.CSS_SELECTOR, '[data-testid="search-box-keyword"]').send_keys(POSITION)
location_input = browser.find_element(By.CSS_SELECTOR, '[data-testid="search-box-location"]')
location_input.send_keys(LOCATION)
try:
    suggestions_container = WebDriverWait(browser, 6).until(
                        expected_conditions.visibility_of_element_located(
                            (By.CSS_SELECTOR, '.absolute.w-full.bg-white.top-full'))
                            )
except TimeoutException as exc:
    raise ValueError("No valid location, try it again.") from exc

suggestions_elements = suggestions_container.find_elements(
                        By.CSS_SELECTOR, 'ul > li > span.text-base.text-blueCorp-600.leading-6')
suggestions_list = [sug_elem.text for sug_elem in suggestions_elements]

MAX_SIMILARITY = 0
MOST_SIMILAR_SUGGESTION = None
for suggestion_element in suggestions_elements:
    suggestion_text = suggestion_element.text
    similarity = difflib.SequenceMatcher(None, LOCATION.lower(), suggestion_text.lower()).ratio()
    if similarity > MAX_SIMILARITY:
        MAX_SIMILARITY = similarity
        MOST_SIMILAR_SUGGESTION = suggestion_element
# Si se encuentra la sugerencia más similar, haz clic en ella
MOST_SIMILAR_SUGGESTION.click()

time.sleep(2)
find_button = browser.find_element(By.CSS_SELECTOR, '[data-testid="search-box-submit"]')
find_button.click()

order_by = WebDriverWait(browser, 5).until(expected_conditions.visibility_of_element_located((By.ID, 'sort-jobs')))
order_by.click()
date_order_button = WebDriverWait(browser, 5).until(expected_conditions.visibility_of_element_located((By.XPATH, '//*[@id="sort-jobs"]/div[2]/div/div[2]/p')))
date_order_button.click()

job_cards = WebDriverWait(browser, 10).until(expected_conditions.presence_of_all_elements_located((By.XPATH, '//*[starts-with(@id, "jobcard-")]')))

# Create dataframe
df = pd.DataFrame()
for card in job_cards:
    card.click()

    # Data extraction
    job_dict = {}
    job_card = WebDriverWait(browser, 7).until(
        expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, '.px-8.py-6.bg-white.rounded-card'))
    )
    title_element = job_card.find_element(By.CSS_SELECTOR, '.flex.justify-between.items-start.mt-2')

    details_elements = WebDriverWait(job_card.find_element(By.CSS_SELECTOR, '.text-lg.mb-4'), 7).until(
        expected_conditions.visibility_of_all_elements_located(
            (By.XPATH,
             "./following-sibling::div[contains(@class, 'mb-1') \
             and not(contains(@class, 'border-t border-[#e4edff]'))]"))
             )
    details_text = "\n".join([elem.text.replace('\n', ' ').replace(',', "\n") for elem in details_elements])

    url_to_apply = browser.current_url

    description_element = job_card.find_element(By.CSS_SELECTOR, '.mb-8.break-words')
    description_text = re.sub(fr'{re.escape("Descripción")}:?', '', description_element.text)

    # Data assignation
    temporal_df = pd.DataFrame({"Title": [title_element.text],
                                "Description": [description_text],
                                "Details": [details_text],
                                "Vacant": [url_to_apply]})

    df = pd.concat([df, temporal_df], ignore_index=True)
df.to_excel("Test.xlsx", sheet_name="test")
# password_input = browser.find_element(by="id").send_keys(USER)
# time.sleep(7)
# # NO OLVIDAR CERRAR EL DRIVER
browser.quit()






#excel = df.to_excel("Test.xlsx", sheet_name="test", index=False)

# Cargar el libro de trabajo y la hoja de cálculo
wb = openpyxl.load_workbook("Test.xlsx")
ws = wb["test"]

# Ajustar automáticamente el ancho de las columnas
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    adjusted_width = (length + 2) * 1.2  # Ajustar el ancho de la columna
    column_index = column_cells[0].column
    column_dimension = openpyxl.worksheet.dimensions.ColumnDimension(
        ws, 
        min=column_index, 
        max=column_index, 
        width=adjusted_width
        )
    ws.column_dimensions[column_index] = column_dimension

# Ajustar altura de filas
for row in ws.iter_rows():
    for cell in row:
        if cell.value:
            lines = str(cell.value).split('\n')
            max_height = max(len(line) for line in lines)
            adjusted_height = max_height * 0.75  # Ajustar la altura de la fila
            cell.alignment = openpyxl.styles.Alignment(wrapText=True)
            # cell.alignment.wrap_text = True
            ws.row_dimensions[cell.row].height = adjusted_height


# Guardar los cambios en el archivo Excel
wb.save("Test.xlsx")

os.system("open Test.xlsx")